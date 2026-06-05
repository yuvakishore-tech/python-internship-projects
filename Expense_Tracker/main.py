import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
import os

# Create folders if not present
os.makedirs("charts", exist_ok=True)
os.makedirs("reports", exist_ok=True)

# Load CSV
df = pd.read_csv("data/expenses.csv")

# Convert Date column
df["Date"] = pd.to_datetime(df["Date"])

# Basic Summary
total_expense = df["Amount"].sum()
highest_expense = df["Amount"].max()
lowest_expense = df["Amount"].min()
average_expense = df["Amount"].mean()

# Category-wise Summary
category_summary = df.groupby("Category")["Amount"].sum()

# ---------------- PIE CHART ----------------
plt.figure(figsize=(7,7))
category_summary.plot(
    kind="pie",
    autopct="%1.1f%%"
)
plt.ylabel("")
plt.title("Expense Distribution by Category")
plt.tight_layout()
plt.savefig("charts/pie.png")
plt.close()

# ---------------- BAR CHART ----------------
plt.figure(figsize=(8,5))
category_summary.plot(kind="bar")
plt.title("Category Wise Expenses")
plt.xlabel("Category")
plt.ylabel("Amount")
plt.tight_layout()
plt.savefig("charts/bar.png")
plt.close()

# ---------------- TREND CHART ----------------
daily_expense = df.groupby("Date")["Amount"].sum()

plt.figure(figsize=(10,5))
daily_expense.plot(kind="line", marker="o")
plt.title("Daily Expense Trend")
plt.xlabel("Date")
plt.ylabel("Amount")
plt.tight_layout()
plt.savefig("charts/trend.png")
plt.close()

# ---------------- EXCEL REPORT ----------------
wb = Workbook()

# Raw Data Sheet
ws1 = wb.active
ws1.title = "Raw_Data"

for col_num, column in enumerate(df.columns, 1):
    cell = ws1.cell(row=1, column=col_num)
    cell.value = column
    cell.font = Font(bold=True)

for row in df.itertuples(index=False):
    ws1.append(list(row))

# Summary Sheet
ws2 = wb.create_sheet(title="Summary")

summary_data = [
    ["Metric", "Value"],
    ["Total Expense", total_expense],
    ["Highest Expense", highest_expense],
    ["Lowest Expense", lowest_expense],
    ["Average Expense", round(average_expense, 2)]
]

for row in summary_data:
    ws2.append(row)

for cell in ws2[1]:
    cell.font = Font(bold=True)

# Category Summary Sheet
ws3 = wb.create_sheet(title="Category_Summary")

ws3.append(["Category", "Total Amount"])

for category, amount in category_summary.items():
    ws3.append([category, amount])

for cell in ws3[1]:
    cell.font = Font(bold=True)

# Add Charts to Excel
img1 = Image("charts/pie.png")
img2 = Image("charts/bar.png")
img3 = Image("charts/trend.png")

ws3.add_image(img1, "D2")
ws3.add_image(img2, "D20")
ws3.add_image(img3, "D38")

# Save Excel Report
report_path = "reports/Expense_Report.xlsx"
wb.save(report_path)

print("Expense Report Generated Successfully")
print("Report saved at:", report_path)