import pandas as pd
import matplotlib.pyplot as plt
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
import os

def generate_report():
    try:
        # Select CSV file
        csv_path = filedialog.askopenfilename(
            title="Select CSV File",
            filetypes=[("CSV Files", "*.csv")]
        )
        if not csv_path:
            return

        df = pd.read_csv(csv_path)
        df.dropna(inplace=True)

        # Summary
        numeric_cols = df.select_dtypes(include="number").columns
        if len(numeric_cols) == 0:
            messagebox.showerror("Error", "No numeric columns found!")
            return

        numeric_col = numeric_cols[0]

        summary_df = pd.DataFrame({
            "Metric": ["Total", "Average", "Minimum", "Maximum"],
            "Value": [
                df[numeric_col].sum(),
                df[numeric_col].mean(),
                df[numeric_col].min(),
                df[numeric_col].max()
            ]
        })

        # Pivot Table 
        pivot_df = pd.pivot_table(
            df,
            index=df.columns[0],
            values=numeric_col,
            aggfunc="sum"
        )

        # Chart
        plt.figure()
        pivot_df.plot(kind="bar", legend=False)
        plt.title("Summary Chart")
        plt.tight_layout()

        chart_path = "chart.png"
        plt.savefig(chart_path)
        plt.close()

        # Save Excel
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",   # ✅ FIXED
            filetypes=[("Excel File", "*.xlsx")]
        )

        if not save_path:
            return

        with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Raw_Data", index=False)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            pivot_df.to_excel(writer, sheet_name="Pivot_Table")

        wb = load_workbook(save_path)

        # Bold headers
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for cell in ws[1]:
                cell.font = Font(bold=True)

        # Add chart
        ws = wb["Pivot_Table"]
        img = Image(chart_path)
        ws.add_image(img, "E2")

        wb.save(save_path)   
        os.remove(chart_path)

        messagebox.showinfo("Success", "Excel Report Generated Successfully")

    except Exception as e:
        messagebox.showerror("Error", str(e))


#  TKINTER UI 
root = tk.Tk()
root.title("Excel Report Generator")
root.geometry("400x200")

label = tk.Label(
    root,
    text="Excel Report Generator\n(Project 8)",
    font=("Arial", 16)
)
label.pack(pady=20)

btn = tk.Button(
    root,
    text="Upload CSV & Generate Report",
    font=("Arial", 12),
    command=generate_report
)
btn.pack(pady=20)

root.mainloop()